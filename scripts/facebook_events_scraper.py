#!/usr/bin/env python3
"""
Facebook Events scraper for musicians, bands, venues, and pages.

Input is a CSV or Excel spreadsheet with one entity per row. Columns are flexible:
use name/source_name/page_name plus optional facebook_url and type. The scraper
logs into Facebook once with --login, saves cookies, then visits public Events
tabs for each entity and extracts both future and past events.

This is designed to integrate with the existing Northeast Ohio live music
pipeline. It writes:
- rows into the shared SQLite events table from neo_live_music_google_scraper.py
- richer rows into a companion facebook_events table
- clean CSV and JSON exports for review/import

It intentionally respects normal browser access. It does not bypass private
groups/pages, solve CAPTCHAs, or access content the logged-in user cannot view.
Keep volume low and use for personal lead discovery.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import json
import logging
import random
import re
import sqlite3
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, urlparse

try:
    from playwright.async_api import Locator, Page
    from playwright.async_api import TimeoutError as PlaywrightTimeoutError
    from playwright.async_api import async_playwright
except Exception:  # pragma: no cover - runtime dependency check gives the fix.
    Locator = Any
    Page = Any
    PlaywrightTimeoutError = None
    async_playwright = None

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from neo_live_music_google_scraper import (  # noqa: E402
    DEFAULT_DB_PATH,
    DEFAULT_EXPORT_DIR,
    EventRecord,
    clean,
    find_location,
    init_db,
    normalize_key,
    parse_event_datetime,
    relevance_score,
    upsert_event,
)


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STORAGE_STATE = (
    Path.home()
    / "Library"
    / "Application Support"
    / "Just Dee Dee Music Map"
    / "facebook-storage-state.json"
)
DEFAULT_OUTPUT_DIR = DEFAULT_EXPORT_DIR / "facebook_events"
LOG_PATH = Path.home() / "Library" / "Logs" / "facebook-events-scraper.log"

FACEBOOK_HOST_RE = re.compile(r"(^|\.)facebook\.com$", re.I)
EVENT_URL_RE = re.compile(r"facebook\.com/(?:events|.+?/events)/([^/?#]+)", re.I)
ADDRESS_RE = re.compile(
    r"\b\d{1,6}\s+[A-Z][A-Za-z0-9 .'-]+?\s+"
    r"(?:St|Street|Rd|Road|Ave|Avenue|Blvd|Boulevard|Dr|Drive|Ln|Lane|Ct|Court|Way|Pkwy|Parkway)\b[^,\n]*",
    re.I,
)


@dataclass
class InputEntity:
    source_name: str
    facebook_url: str = ""
    entity_type: str = ""
    row_number: int = 0


@dataclass
class FacebookEvent:
    event_url: str
    event_id: str
    event_title: str
    organizer_name: str
    page_name: str
    start_datetime: str
    end_datetime: str
    venue: str
    location: str
    address: str
    description: str
    interested_count: int
    going_count: int
    ticket_url: str
    price_info: str
    image_urls: str
    event_type: str
    scraped_at: str
    source_name: str
    source_url: str
    raw_text: str

    @property
    def dedupe_fingerprint(self) -> str:
        return stable_hash(normalize_key(self.event_title), self.start_datetime, normalize_key(self.organizer_name or self.page_name), length=24)


def stable_hash(*parts: object, length: int = 16) -> str:
    return hashlib.sha256("|".join(clean(part) for part in parts).encode("utf-8")).hexdigest()[:length]


def parse_count(value: object) -> int:
    text = clean(value).replace(",", "")
    match = re.search(r"(\d+(?:\.\d+)?)\s*([KkMm]?)", text)
    if not match:
        return 0
    number = float(match.group(1))
    suffix = match.group(2).lower()
    if suffix == "k":
        number *= 1000
    elif suffix == "m":
        number *= 1_000_000
    return int(number)


def extract_counts(text: str) -> tuple[int, int]:
    interested = 0
    going = 0
    interested_match = re.search(r"(\d+(?:,\d{3})*(?:\.\d+)?\s*[KkMm]?)\s+(?:interested|are interested)", text, re.I)
    going_match = re.search(r"(\d+(?:,\d{3})*(?:\.\d+)?\s*[KkMm]?)\s+(?:going|are going)", text, re.I)
    if interested_match:
        interested = parse_count(interested_match.group(1))
    if going_match:
        going = parse_count(going_match.group(1))
    return interested, going


def parse_datetime_candidate(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    candidate = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(candidate).isoformat(timespec="seconds")
    except ValueError:
        pass
    event_date, event_time = parse_event_datetime(text)
    if not event_date:
        return ""
    if event_time:
        try:
            parsed_time = datetime.strptime(event_time, "%I:%M %p").time()
            return datetime.combine(date.fromisoformat(event_date), parsed_time).isoformat(timespec="seconds")
        except ValueError:
            return event_date
    return event_date


def parse_facebook_event_datetimes(text: str) -> tuple[str, str]:
    """Parse common Facebook event date strings visible in 2026 layouts."""
    text = clean(text)
    patterns = [
        r"\b(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun),?\s+([A-Z][a-z]+)\s+(\d{1,2}),?\s*(\d{4})?\s+(?:at\s+)?(\d{1,2}(?::\d{2})?\s*[AP]M)",
        r"\b([A-Z][a-z]+)\s+(\d{1,2}),?\s*(\d{4})?\s+(?:at\s+)?(\d{1,2}(?::\d{2})?\s*[AP]M)",
        r"\b([A-Z][a-z]+)\s+(\d{1,2})\s*-\s*([A-Z][a-z]+)?\s*(\d{1,2}),?\s*(\d{4})",
    ]
    today = date.today()
    for pattern in patterns[:2]:
        match = re.search(pattern, text)
        if not match:
            continue
        month_text, day_text, year_text, time_text = match.groups()
        try:
            month = datetime.strptime(month_text[:3].title(), "%b").month
            year = int(year_text or today.year)
            parsed_time = parse_time_text(time_text)
            if parsed_time:
                return datetime(year, month, int(day_text), parsed_time.hour, parsed_time.minute).isoformat(timespec="seconds"), ""
            return date(year, month, int(day_text)).isoformat(), ""
        except ValueError:
            continue
    range_match = re.search(patterns[2], text)
    if range_match:
        start_month_text, start_day, end_month_text, end_day, year_text = range_match.groups()
        try:
            start_month = datetime.strptime(start_month_text[:3].title(), "%b").month
            end_month = datetime.strptime((end_month_text or start_month_text)[:3].title(), "%b").month
            year = int(year_text)
            return date(year, start_month, int(start_day)).isoformat(), date(year, end_month, int(end_day)).isoformat()
        except ValueError:
            pass
    start_date, start_time = parse_event_datetime(text)
    if start_date and start_time:
        parsed_time = parse_time_text(start_time)
        if parsed_time:
            return datetime.combine(date.fromisoformat(start_date), parsed_time).isoformat(timespec="seconds"), ""
    return start_date, ""


def parse_time_text(value: object) -> datetime.time | None:
    match = re.search(r"\b(\d{1,2})(?::(\d{2}))?\s*([AP]M)\b", clean(value), re.I)
    if not match:
        return None
    hour = int(match.group(1)) % 12
    if match.group(3).upper() == "PM":
        hour += 12
    return datetime.strptime(f"{hour:02d}:{int(match.group(2) or 0):02d}", "%H:%M").time()


def event_id_from_url(url: str, fallback: str = "") -> str:
    match = EVENT_URL_RE.search(url)
    if match:
        return clean(match.group(1)).strip("/")
    query = parse_qs(urlparse(url).query)
    for key in ("id", "event_id"):
        if query.get(key):
            return clean(query[key][0])
    return stable_hash(url, fallback, length=18)


def normalize_facebook_url(url: str) -> str:
    url = clean(url)
    if not url:
        return ""
    if not re.match(r"https?://", url):
        url = "https://" + url.lstrip("/")
    parsed = urlparse(url)
    if "m.facebook.com" in parsed.netloc:
        url = url.replace("m.facebook.com", "www.facebook.com")
    return url.split("#")[0]


def events_url_for_page(url: str) -> str:
    url = normalize_facebook_url(url).rstrip("/")
    if not url:
        return ""
    parsed = urlparse(url)
    if "/events/" in parsed.path and re.search(r"/events/[^/]+", parsed.path):
        return url
    if parsed.path.startswith("/groups/"):
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) >= 2:
            return f"https://www.facebook.com/groups/{parts[1]}/events"
    if parsed.path == "/profile.php":
        separator = "&" if parsed.query else "?"
        return f"{url}{separator}sk=events"
    if parsed.path.endswith("/events"):
        return url
    return f"{url}/events"


def is_event_in_mode(event: FacebookEvent, mode: str) -> bool:
    if mode == "all":
        return True
    start = parse_date_from_datetime(event.start_datetime)
    if not start:
        return True
    return start >= date.today()


def parse_date_from_datetime(value: str) -> date | None:
    text = clean(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def load_spreadsheet(path: Path) -> list[InputEntity]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return load_csv_entities(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_excel_entities(path)
    raise ValueError(f"Unsupported input file type: {path.suffix}. Use CSV, TSV, XLSX, or XLSM.")


def load_csv_entities(path: Path) -> list[InputEntity]:
    sample = path.read_text(encoding="utf-8-sig")[:4096]
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    rows = list(csv.reader(path.read_text(encoding="utf-8-sig").splitlines(), delimiter=delimiter))
    if not rows:
        return []
    first = [normalize_key(cell) for cell in rows[0]]
    has_header = any(header in {"name", "source name", "source_name", "page name", "page_name", "facebook url", "facebook_url", "url", "type"} for header in first)
    entities: list[InputEntity] = []
    if has_header:
        headers = [normalize_key(cell).replace(" ", "_") for cell in rows[0]]
        for row_number, row in enumerate(rows[1:], start=2):
            data = {headers[index]: clean(value) for index, value in enumerate(row) if index < len(headers)}
            entity = entity_from_row(data, row_number)
            if entity:
                entities.append(entity)
    else:
        for row_number, row in enumerate(rows, start=1):
            if row and clean(row[0]):
                entities.append(InputEntity(source_name=clean(row[0]), facebook_url=clean(row[1]) if len(row) > 1 else "", row_number=row_number))
    return entities


def load_excel_entities(path: Path) -> list[InputEntity]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Excel input requires openpyxl. Run: python3 -m pip install -r requirements-scraper.txt") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return []
    first = [normalize_key(cell) for cell in rows[0]]
    has_header = any(header in {"name", "source name", "source_name", "page name", "page_name", "facebook url", "facebook_url", "url", "type"} for header in first)
    entities: list[InputEntity] = []
    if has_header:
        headers = [normalize_key(cell).replace(" ", "_") for cell in rows[0]]
        for row_number, row in enumerate(rows[1:], start=2):
            data = {headers[index]: clean(value) for index, value in enumerate(row) if index < len(headers)}
            entity = entity_from_row(data, row_number)
            if entity:
                entities.append(entity)
    else:
        for row_number, row in enumerate(rows, start=1):
            if row and clean(row[0]):
                entities.append(InputEntity(source_name=clean(row[0]), facebook_url=clean(row[1]) if len(row) > 1 else "", row_number=row_number))
    return entities


def entity_from_row(data: dict[str, str], row_number: int) -> InputEntity | None:
    name = (
        data.get("source_name")
        or data.get("name")
        or data.get("page_name")
        or data.get("artist")
        or data.get("band")
        or data.get("venue")
    )
    facebook_url = data.get("facebook_url") or data.get("facebook") or data.get("url") or ""
    entity_type = data.get("type") or data.get("entity_type") or ""
    if not clean(name) and not clean(facebook_url):
        return None
    return InputEntity(source_name=clean(name) or clean(facebook_url), facebook_url=normalize_facebook_url(facebook_url), entity_type=clean(entity_type), row_number=row_number)


def init_facebook_events_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS facebook_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_url TEXT NOT NULL,
            event_id TEXT,
            event_title TEXT,
            organizer_name TEXT,
            page_name TEXT,
            start_datetime TEXT,
            end_datetime TEXT,
            venue TEXT,
            location TEXT,
            address TEXT,
            description TEXT,
            interested_count INTEGER DEFAULT 0,
            going_count INTEGER DEFAULT 0,
            ticket_url TEXT,
            price_info TEXT,
            image_urls TEXT,
            event_type TEXT,
            scraped_at TEXT NOT NULL,
            source_name TEXT,
            source_url TEXT,
            raw_text TEXT,
            dedupe_fingerprint TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_facebook_events_url ON facebook_events(event_url)")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_facebook_events_fingerprint ON facebook_events(dedupe_fingerprint)")
    conn.commit()


def upsert_facebook_event(conn: sqlite3.Connection, event: FacebookEvent) -> None:
    row = {**asdict(event), "dedupe_fingerprint": event.dedupe_fingerprint}
    try:
        conn.execute(
            """
            INSERT INTO facebook_events (
                event_url, event_id, event_title, organizer_name, page_name,
                start_datetime, end_datetime, venue, location, address,
                description, interested_count, going_count, ticket_url,
                price_info, image_urls, event_type, scraped_at, source_name,
                source_url, raw_text, dedupe_fingerprint
            ) VALUES (
                :event_url, :event_id, :event_title, :organizer_name, :page_name,
                :start_datetime, :end_datetime, :venue, :location, :address,
                :description, :interested_count, :going_count, :ticket_url,
                :price_info, :image_urls, :event_type, :scraped_at, :source_name,
                :source_url, :raw_text, :dedupe_fingerprint
            )
            ON CONFLICT(event_url) DO UPDATE SET
                event_title=excluded.event_title,
                organizer_name=excluded.organizer_name,
                page_name=excluded.page_name,
                start_datetime=excluded.start_datetime,
                end_datetime=excluded.end_datetime,
                venue=excluded.venue,
                location=excluded.location,
                address=excluded.address,
                description=excluded.description,
                interested_count=excluded.interested_count,
                going_count=excluded.going_count,
                ticket_url=excluded.ticket_url,
                price_info=excluded.price_info,
                image_urls=excluded.image_urls,
                event_type=excluded.event_type,
                scraped_at=excluded.scraped_at,
                source_name=excluded.source_name,
                source_url=excluded.source_url,
                raw_text=excluded.raw_text,
                dedupe_fingerprint=excluded.dedupe_fingerprint,
                updated_at=CURRENT_TIMESTAMP
            """,
            row,
        )
    except sqlite3.IntegrityError:
        conn.execute(
            """
            UPDATE facebook_events SET
                event_url=:event_url,
                event_id=:event_id,
                event_title=:event_title,
                organizer_name=:organizer_name,
                page_name=:page_name,
                start_datetime=:start_datetime,
                end_datetime=:end_datetime,
                venue=:venue,
                location=:location,
                address=:address,
                description=:description,
                interested_count=:interested_count,
                going_count=:going_count,
                ticket_url=:ticket_url,
                price_info=:price_info,
                image_urls=:image_urls,
                event_type=:event_type,
                scraped_at=:scraped_at,
                source_name=:source_name,
                source_url=:source_url,
                raw_text=:raw_text,
                updated_at=CURRENT_TIMESTAMP
            WHERE dedupe_fingerprint=:dedupe_fingerprint
            """,
            row,
        )
    conn.commit()


def facebook_event_to_event_record(event: FacebookEvent) -> EventRecord:
    event_date = ""
    event_time = ""
    if event.start_datetime:
        event_date = event.start_datetime[:10]
        if "T" in event.start_datetime:
            try:
                parsed = datetime.fromisoformat(event.start_datetime)
                event_time = parsed.strftime("%-I:%M %p") if sys.platform != "win32" else parsed.strftime("%I:%M %p").lstrip("0")
            except ValueError:
                event_time = ""
    city, distance = find_location(" ".join([event.location, event.address, event.venue, event.description]))
    return EventRecord(
        source="facebook_events",
        query_used=event.source_name,
        title=event.event_title,
        url=event.event_url,
        event_date=event_date,
        event_time=event_time,
        venue=event.venue,
        address=event.address,
        city=city or event.location,
        bands=event.organizer_name or event.page_name,
        ticket_info=event.ticket_url or event.price_info,
        description=event.description[:1200],
        image_urls=event.image_urls,
        scraped_at=event.scraped_at,
        raw_snippet=event.raw_text[:1200],
        location_match=city,
        distance_miles=round(distance, 1) if distance is not None else None,
        relevance_score=max(8, relevance_score(" ".join([event.event_title, event.description, event.venue, event.location]))),
    )


def upsert_event_record(conn: sqlite3.Connection, event_record: EventRecord) -> None:
    try:
        upsert_event(conn, event_record)
        return
    except sqlite3.IntegrityError:
        title_date_key = hashlib.sha256(
            f"{normalize_key(event_record.title)}|{event_record.event_date or 'undated'}".encode("utf-8")
        ).hexdigest()
        conn.execute(
            """
            UPDATE events SET
                source=:source,
                query_used=:query_used,
                url=:url,
                event_time=:event_time,
                venue=:venue,
                address=:address,
                city=:city,
                bands=:bands,
                ticket_info=:ticket_info,
                description=:description,
                image_urls=:image_urls,
                scraped_at=:scraped_at,
                raw_snippet=:raw_snippet,
                location_match=:location_match,
                distance_miles=:distance_miles,
                relevance_score=:relevance_score,
                updated_at=CURRENT_TIMESTAMP
            WHERE title_date_key=:title_date_key
            """,
            {**asdict(event_record), "title_date_key": title_date_key},
        )
        conn.commit()


async def random_delay(low: float = 1.0, high: float = 4.0) -> None:
    await asyncio.sleep(random.uniform(low, high))


async def slow_mouse_wiggle(page: Page) -> None:
    try:
        width = random.randint(900, 1300)
        height = random.randint(500, 850)
        for _ in range(random.randint(2, 4)):
            await page.mouse.move(random.randint(40, width), random.randint(40, height), steps=random.randint(8, 18))
            await asyncio.sleep(random.uniform(0.1, 0.35))
    except Exception:
        pass


async def click_text_if_visible(page_or_locator: Page | Locator, patterns: list[str], timeout_ms: int = 1200) -> bool:
    for pattern in patterns:
        try:
            locator = page_or_locator.get_by_text(re.compile(pattern, re.I)).first
            if await locator.count():
                await locator.click(timeout=timeout_ms)
                await random_delay(0.4, 1.2)
                return True
        except Exception:
            continue
    return False


async def expand_visible_text(page: Page) -> None:
    for _ in range(6):
        clicked = await click_text_if_visible(page, [r"^See more$", r"^More$"], timeout_ms=800)
        if not clicked:
            break


async def click_load_more(page: Page) -> bool:
    return await click_text_if_visible(
        page,
        [
            r"^See more events$",
            r"^Load more$",
            r"^See more$",
            r"^Show more$",
            r"^More events$",
        ],
        timeout_ms=1800,
    )


async def robust_goto(page: Page, url: str, timeout_ms: int, logger: logging.Logger, retries: int = 2) -> bool:
    for attempt in range(retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            await random_delay(1.0, 2.6)
            return True
        except Exception as exc:
            logger.warning("Navigation failed %s attempt %d/%d: %s", url, attempt + 1, retries + 1, exc)
            await random_delay(2.0, 5.0)
    return False


class FacebookEventsScraper:
    def __init__(self, args: argparse.Namespace):
        self.args = args
        self.logger = logging.getLogger("facebook_events")
        self.page: Page | None = None

    async def run(self, entities: list[InputEntity]) -> list[FacebookEvent]:
        if async_playwright is None:
            raise RuntimeError("Playwright is not installed. Run: python3 -m pip install -r requirements-scraper.txt && python3 -m playwright install chromium")
        all_events: list[FacebookEvent] = []
        async with async_playwright() as playwright:
            launch_options: dict[str, Any] = {"headless": self.args.headless, "slow_mo": self.args.slow_mo}
            if self.args.browser_channel:
                launch_options["channel"] = self.args.browser_channel
            if self.args.proxy_server:
                launch_options["proxy"] = {"server": self.args.proxy_server}
            browser = await playwright.chromium.launch(**launch_options)
            context_options: dict[str, Any] = {
                "locale": "en-US",
                "timezone_id": "America/New_York",
                "viewport": {"width": random.randint(1280, 1480), "height": random.randint(820, 980)},
                "user_agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
            }
            storage_state = Path(self.args.storage_state).expanduser()
            if storage_state.exists():
                context_options["storage_state"] = str(storage_state)
            context = await browser.new_context(**context_options)
            self.page = await context.new_page()
            for entity in entities:
                try:
                    entity_events = await self.process_entity(entity)
                    self.logger.info("Processing %s - %d events found", entity.source_name, len(entity_events))
                    all_events.extend(entity_events)
                except Exception as exc:
                    self.logger.exception("Skipping failed entity %s row %s: %s", entity.source_name, entity.row_number, exc)
                await random_delay(2.0, 6.0)
            await context.close()
            await browser.close()
        return all_events

    async def process_entity(self, entity: InputEntity) -> list[FacebookEvent]:
        page_url = normalize_facebook_url(entity.facebook_url)
        if not page_url:
            page_url = await self.search_page(entity.source_name, entity.entity_type)
        if not page_url:
            self.logger.warning("No Facebook page/profile found for %s", entity.source_name)
            return []
        events = await self.scrape_events_for_page(page_url, entity.source_name)
        return [event for event in events if is_event_in_mode(event, self.args.mode)]

    async def search_page(self, name: str, entity_type: str = "") -> str:
        if self.page is None:
            return ""
        search_kinds = ["pages"]
        if normalize_key(entity_type) in {"person", "profile", "musician", "artist"}:
            search_kinds = ["people", "pages"]
        for search_kind in search_kinds:
            search_url = f"https://www.facebook.com/search/{search_kind}/?q={quote_plus(name)}"
            ok = await robust_goto(self.page, search_url, self.args.timeout_ms, self.logger)
            if not ok or await self.is_login_wall():
                return ""
            await slow_mouse_wiggle(self.page)
            links = await self.page.evaluate(
                """
                (needle) => Array.from(document.querySelectorAll('a'))
                    .map(a => ({
                        href: a.href || '',
                        text: (a.innerText || a.textContent || '').replace(/\\s+/g, ' ').trim(),
                        aria: (a.getAttribute('aria-label') || '').replace(/\\s+/g, ' ').trim()
                    }))
                    .filter(item => item.href.includes('facebook.com') && !/\\/search\\/|\\/login|\\/share|\\/help\\//.test(item.href))
                    .map(item => ({...item, score:
                        ((item.text || item.aria).toLowerCase().includes(needle.toLowerCase()) ? 10 : 0) +
                        (/\\/pages\\//.test(item.href) ? 2 : 0)
                    }))
                    .sort((a, b) => b.score - a.score)
                    .slice(0, 15)
                """,
                name,
            )
            for item in links:
                href = normalize_facebook_url(clean(item.get("href")).split("?")[0])
                if href and FACEBOOK_HOST_RE.search(urlparse(href).netloc):
                    return href
        return ""

    async def scrape_events_for_page(self, page_url: str, source_name: str) -> list[FacebookEvent]:
        if self.page is None:
            return []
        if EVENT_URL_RE.search(page_url):
            event = await self.scrape_event_detail(page_url, source_name, page_url)
            return [event] if event else []
        candidate_urls = self.events_tab_urls(page_url)
        event_links: dict[str, str] = {}
        for url in candidate_urls:
            ok = await robust_goto(self.page, url, self.args.timeout_ms, self.logger)
            if not ok:
                continue
            if await self.is_login_wall():
                self.logger.warning("Login/restriction wall on %s", url)
                continue
            await self.try_events_tabs()
            mode_scrolls = self.args.max_scrolls if self.args.mode == "all" else max(3, self.args.max_scrolls // 2)
            for _ in range(mode_scrolls):
                await expand_visible_text(self.page)
                await slow_mouse_wiggle(self.page)
                links = await self.collect_event_links(source_name)
                event_links.update(links)
                clicked = await click_load_more(self.page)
                await self.page.mouse.wheel(0, random.randint(1000, 1800))
                await random_delay(1.0, 3.2)
                if not clicked and len(event_links) >= self.args.max_events_per_entity:
                    break
            if len(event_links) >= self.args.max_events_per_entity:
                break
        events: list[FacebookEvent] = []
        for event_url in list(event_links.keys())[: self.args.max_events_per_entity]:
            event = await self.scrape_event_detail(event_url, source_name, page_url)
            if event and is_event_in_mode(event, self.args.mode):
                events.append(event)
            await random_delay(1.2, 4.0)
        return dedupe_events(events)

    def events_tab_urls(self, page_url: str) -> list[str]:
        base_events_url = events_url_for_page(page_url)
        root = normalize_facebook_url(page_url).rstrip("/")
        urls = [base_events_url]
        if "/events/" not in root and not root.endswith("/events"):
            urls.extend([f"{root}/events_hosted", f"{root}/upcoming_hosted_events"])
        if self.args.mode == "all":
            urls.extend(
                [
                    f"{base_events_url}?active_tab=about",
                    f"{base_events_url}?active_tab=past",
                    f"{base_events_url}?sk=past_hosted_events",
                    f"{root}/past_hosted_events",
                ]
            )
        return list(dict.fromkeys(urls))

    async def try_events_tabs(self) -> None:
        if self.page is None:
            return
        labels = ["Upcoming", "Upcoming events", "Future events"]
        if self.args.mode == "all":
            labels.extend(["Past", "Past events", "All events", "Events"])
        for label in labels:
            try:
                locator = self.page.get_by_role("tab", name=re.compile(label, re.I)).first
                if await locator.count():
                    await locator.click(timeout=1200)
                    await random_delay(1.0, 2.0)
            except Exception:
                try:
                    await self.page.get_by_text(re.compile(rf"^{re.escape(label)}$", re.I)).first.click(timeout=1000)
                    await random_delay(1.0, 2.0)
                except Exception:
                    continue

    async def collect_event_links(self, source_name: str) -> dict[str, str]:
        if self.page is None:
            return {}
        raw_links = await self.page.evaluate(
            """
            () => Array.from(document.querySelectorAll('a'))
                .map(a => ({
                    href: a.href || '',
                    text: (a.innerText || a.textContent || '').replace(/\\s+/g, ' ').trim()
                }))
                .filter(item => /facebook\\.com\\/events\\//.test(item.href))
                .slice(0, 300)
            """
        )
        links: dict[str, str] = {}
        for item in raw_links:
            href = normalize_facebook_url(clean(item.get("href")).split("?")[0])
            if not href or "/events/create" in href or href.endswith("/events"):
                continue
            event_id = event_id_from_url(href, item.get("text"))
            if event_id:
                links[href] = clean(item.get("text")) or source_name
        return links

    async def scrape_event_detail(self, event_url: str, source_name: str, source_url: str) -> FacebookEvent | None:
        if self.page is None:
            return None
        ok = await robust_goto(self.page, event_url, self.args.timeout_ms, self.logger)
        if not ok or await self.is_login_wall():
            self.logger.warning("Restricted or unavailable Facebook event: %s", event_url)
            return None
        await expand_visible_text(self.page)
        for _ in range(2):
            await self.page.mouse.wheel(0, random.randint(600, 1300))
            await random_delay(0.8, 1.8)
            await expand_visible_text(self.page)
        raw = await self.extract_event_page_data()
        return build_event_from_page_data(raw, event_url, source_name, source_url)

    async def extract_event_page_data(self) -> dict[str, Any]:
        if self.page is None:
            return {}
        return await self.page.evaluate(
            """
            () => {
                const clean = value => (value || '').replace(/\\s+/g, ' ').trim();
                const meta = selector => document.querySelector(selector)?.content || '';
                const links = Array.from(document.querySelectorAll('a')).map(a => ({
                    href: a.href || '',
                    text: clean(a.innerText || a.textContent || ''),
                    aria: clean(a.getAttribute('aria-label') || '')
                })).filter(item => item.href);
                const ticketLinks = links.filter(item =>
                    /ticket|eventbrite|bandsintown|dice\\.fm|showclix|seetickets|ticketmaster/i.test(
                        [item.href, item.text, item.aria].join(' ')
                    )
                );
                const pageLinks = links.filter(item =>
                    /facebook\\.com\\//.test(item.href) &&
                    !/\\/events\\/|\\/search\\/|\\/share|\\/photo|\\/login/.test(item.href)
                );
                const images = Array.from(document.querySelectorAll('img'))
                    .map(img => img.currentSrc || img.src || '')
                    .filter(src => /^https?:/.test(src));
                const bodyText = document.body ? (document.body.innerText || '') : '';
                const jsonld = Array.from(document.querySelectorAll('script[type="application/ld+json"]'))
                    .map(script => script.textContent || '');
                return {
                    url: location.href,
                    title: clean(document.querySelector('h1')?.innerText || meta('meta[property="og:title"]') || document.title),
                    metaTitle: clean(meta('meta[property="og:title"]')),
                    metaDescription: clean(meta('meta[property="og:description"]') || meta('meta[name="description"]')),
                    metaImage: clean(meta('meta[property="og:image"]')),
                    bodyText,
                    ticketLinks,
                    pageLinks,
                    imageUrls: Array.from(new Set([meta('meta[property="og:image"]'), ...images].filter(Boolean))),
                    jsonld
                };
            }
            """
        )

    async def is_login_wall(self) -> bool:
        if self.page is None:
            return False
        text = clean(await self.page.evaluate("document.body ? document.body.innerText : ''")).lower()
        return "log into facebook" in text or "you must log in" in text or "create new account" in text


def build_event_from_page_data(raw: dict[str, Any], event_url: str, source_name: str, source_url: str) -> FacebookEvent:
    jsonld = parse_jsonld(raw.get("jsonld") or [])
    title = clean(raw.get("title")) or clean(raw.get("metaTitle")) or "Facebook Event"
    title = re.sub(r"\s*\|\s*Facebook\s*$", "", title, flags=re.I)
    body_text = clean(raw.get("bodyText"))
    meta_description = clean(raw.get("metaDescription"))
    description = choose_description(body_text, meta_description)
    start_datetime, end_datetime = pick_datetimes(jsonld, body_text, meta_description)
    venue, location, address = pick_location(jsonld, body_text)
    organizer = pick_organizer(jsonld, raw.get("pageLinks") or [], source_name)
    interested_count, going_count = extract_counts(body_text)
    ticket_url, price_info = pick_ticket(raw.get("ticketLinks") or [], body_text, jsonld)
    image_urls = " | ".join(list(dict.fromkeys(clean(url) for url in raw.get("imageUrls") or [] if clean(url).startswith("http")))[:12])
    event_type = "online" if re.search(r"\bonline event\b|facebook live|virtual", body_text, re.I) else "in-person"
    return FacebookEvent(
        event_url=normalize_facebook_url(event_url),
        event_id=event_id_from_url(event_url, title),
        event_title=title,
        organizer_name=organizer,
        page_name=organizer,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        venue=venue,
        location=location,
        address=address,
        description=description,
        interested_count=interested_count,
        going_count=going_count,
        ticket_url=ticket_url,
        price_info=price_info,
        image_urls=image_urls,
        event_type=event_type,
        scraped_at=datetime.now().isoformat(timespec="seconds"),
        source_name=source_name,
        source_url=source_url,
        raw_text=body_text,
    )


def parse_jsonld(raw_items: list[str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for item in raw_items:
        try:
            parsed = json.loads(item)
        except Exception:
            continue
        if isinstance(parsed, list):
            records.extend(record for record in parsed if isinstance(record, dict))
        elif isinstance(parsed, dict):
            graph = parsed.get("@graph")
            if isinstance(graph, list):
                records.extend(record for record in graph if isinstance(record, dict))
            records.append(parsed)
    return records


def choose_description(body_text: str, meta_description: str) -> str:
    if meta_description and len(meta_description) > 40:
        return meta_description
    lines = [line.strip() for line in re.split(r"\n| {2,}", body_text) if clean(line)]
    stop = {"About", "Discussion", "Invite", "Share", "Details", "Event by", "Public", "Anyone on or off Facebook"}
    useful = [line for line in lines if line not in stop and not re.fullmatch(r"\d+[KkMm]?", line)]
    return clean("\n".join(useful[:80]))


def pick_datetimes(jsonld: list[dict[str, Any]], body_text: str, meta_description: str) -> tuple[str, str]:
    for record in jsonld:
        if "startDate" in record:
            start = parse_datetime_candidate(record.get("startDate"))
            end = parse_datetime_candidate(record.get("endDate"))
            if start:
                return start, end
    start, end = parse_facebook_event_datetimes(" ".join([meta_description, body_text[:3000]]))
    return start, end


def pick_location(jsonld: list[dict[str, Any]], body_text: str) -> tuple[str, str, str]:
    for record in jsonld:
        location = record.get("location")
        if isinstance(location, dict):
            name = clean(location.get("name"))
            address_obj = location.get("address")
            if isinstance(address_obj, dict):
                parts = [
                    clean(address_obj.get("streetAddress")),
                    clean(address_obj.get("addressLocality")),
                    clean(address_obj.get("addressRegion")),
                    clean(address_obj.get("postalCode")),
                ]
                address = clean(", ".join(part for part in parts if part))
                city = clean(address_obj.get("addressLocality"))
                return name, city, address
            if isinstance(address_obj, str):
                city, _ = find_location(address_obj)
                return name, city, clean(address_obj)
    address_match = ADDRESS_RE.search(body_text)
    address = clean(address_match.group(0)) if address_match else ""
    city, _ = find_location(body_text)
    venue = ""
    lines = [clean(line) for line in body_text.splitlines() if clean(line)]
    for index, line in enumerate(lines):
        if line.lower() in {"location", "where"} and index + 1 < len(lines):
            venue = lines[index + 1]
            break
    return venue, city, address


def pick_organizer(jsonld: list[dict[str, Any]], page_links: list[dict[str, str]], source_name: str) -> str:
    for record in jsonld:
        organizer = record.get("organizer") or record.get("performer")
        if isinstance(organizer, dict) and clean(organizer.get("name")):
            return clean(organizer.get("name"))
        if isinstance(organizer, list):
            names = [clean(item.get("name")) for item in organizer if isinstance(item, dict) and clean(item.get("name"))]
            if names:
                return names[0]
    for item in page_links:
        text = clean(item.get("text") or item.get("aria"))
        if text and not re.search(r"facebook|home|events|photos|videos|about|more", text, re.I):
            return text
    return source_name


def pick_ticket(ticket_links: list[dict[str, str]], body_text: str, jsonld: list[dict[str, Any]]) -> tuple[str, str]:
    for record in jsonld:
        offers = record.get("offers")
        if isinstance(offers, dict):
            price = clean(offers.get("price"))
            url = clean(offers.get("url"))
            if price or url:
                return url, f"${price}" if price else ""
    for item in ticket_links:
        href = clean(item.get("href"))
        if href:
            return href, clean(item.get("text") or item.get("aria"))
    price_match = re.search(r"\b(?:free|cover[:\s]+.{1,40}|tickets?[:\s]+.{1,80}|\$\d+(?:\.\d{2})?)", body_text, re.I)
    return "", clean(price_match.group(0)) if price_match else ""


def dedupe_events(events: list[FacebookEvent]) -> list[FacebookEvent]:
    by_key: dict[str, FacebookEvent] = {}
    for event in events:
        key = event.event_url or event.dedupe_fingerprint
        if key not in by_key:
            by_key[key] = event
    return list(by_key.values())


async def login(storage_state: Path, browser_channel: str = "") -> None:
    if async_playwright is None:
        raise RuntimeError("Playwright is not installed. Run: python3 -m pip install -r requirements-scraper.txt && python3 -m playwright install chromium")
    storage_state = storage_state.expanduser()
    storage_state.parent.mkdir(parents=True, exist_ok=True)
    async with async_playwright() as playwright:
        launch_options: dict[str, Any] = {"headless": False}
        if browser_channel:
            launch_options["channel"] = browser_channel
        browser = await playwright.chromium.launch(**launch_options)
        context_options: dict[str, Any] = {
            "locale": "en-US",
            "timezone_id": "America/New_York",
            "viewport": {"width": 1360, "height": 920},
        }
        if storage_state.exists():
            context_options["storage_state"] = str(storage_state)
        context = await browser.new_context(**context_options)
        page = await context.new_page()
        await page.goto("https://www.facebook.com/login", wait_until="domcontentloaded", timeout=60000)
        print("Facebook login window is open. Log in normally, then press Enter here to save the session.")
        await asyncio.to_thread(input)
        await context.storage_state(path=str(storage_state))
        await context.close()
        await browser.close()
    print(f"Saved Facebook session to {storage_state}")


def export_outputs(events: list[FacebookEvent], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"facebook_events_{timestamp}.csv"
    json_path = output_dir / f"facebook_events_{timestamp}.json"
    latest_csv = output_dir / "facebook_events_latest.csv"
    latest_json = output_dir / "facebook_events_latest.json"
    fields = [field.name for field in FacebookEvent.__dataclass_fields__.values()]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for event in events:
            writer.writerow(asdict(event))
    json_path.write_text(json.dumps([asdict(event) for event in events], indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    latest_csv.write_text(csv_path.read_text(encoding="utf-8"), encoding="utf-8")
    latest_json.write_text(json_path.read_text(encoding="utf-8"), encoding="utf-8")
    return {"csv": csv_path, "json": json_path, "latest_csv": latest_csv, "latest_json": latest_json}


def save_events(db_path: Path, events: list[FacebookEvent]) -> None:
    conn = init_db(db_path)
    init_facebook_events_table(conn)
    for event in events:
        upsert_facebook_event(conn, event)
        upsert_event_record(conn, facebook_event_to_event_record(event))
    conn.close()


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Facebook Events for a spreadsheet of musicians, bands, venues, or pages.")
    parser.add_argument("--input", type=Path, help="CSV/XLSX with one source per row and optional facebook_url/type columns.")
    parser.add_argument("--login", action="store_true", help="Open Facebook headful and save a reusable login session.")
    parser.add_argument("--storage-state", type=Path, default=DEFAULT_STORAGE_STATE)
    parser.add_argument("--limit", "--max-events-per-entity", dest="max_events_per_entity", type=int, default=50)
    parser.add_argument("--mode", choices=["future_only", "all"], default="all")
    parser.add_argument("--headless", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--timeout-ms", type=int, default=45000)
    parser.add_argument("--max-scrolls", type=int, default=10)
    parser.add_argument("--slow-mo", type=int, default=0, help="Playwright slow_mo in milliseconds for visible debugging.")
    parser.add_argument("--db-path", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--proxy-server", default="", help="Optional Playwright proxy, e.g. http://host:port")
    parser.add_argument("--browser-channel", default="", help="Optional Playwright browser channel, e.g. chrome")
    parser.add_argument("--log-level", default="INFO")
    return parser.parse_args(argv)


def configure_logging(level: str) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if sys.stdout.isatty():
        handlers.append(logging.FileHandler(LOG_PATH, encoding="utf-8"))
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
        handlers=handlers,
        force=True,
    )


async def async_main(argv: list[str]) -> int:
    args = parse_args(argv)
    configure_logging(args.log_level)
    logger = logging.getLogger("facebook_events")
    if args.login:
        await login(args.storage_state, args.browser_channel)
        return 0
    if not args.input:
        raise SystemExit("--input is required unless --login is used.")
    entities = load_spreadsheet(args.input)
    logger.info("Loaded %d input entities from %s", len(entities), args.input)
    scraper = FacebookEventsScraper(args)
    events = await scraper.run(entities)
    events = dedupe_events(events)
    save_events(args.db_path, events)
    paths = export_outputs(events, args.output_dir)
    summary = {
        "entities": len(entities),
        "events": len(events),
        "db_path": str(args.db_path),
        "outputs": {key: str(value) for key, value in paths.items()},
    }
    logger.info("Saved %d Facebook Events to %s", len(events), args.db_path)
    print(json.dumps({"summary": summary, "events": [asdict(event) for event in events]}, indent=2, ensure_ascii=False))
    return 0


def main(argv: list[str]) -> int:
    return asyncio.run(async_main(argv))


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
